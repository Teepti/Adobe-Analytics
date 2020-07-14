import csv
import openpyxl
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import os
from pathlib import Path


def AnalyticsValidation():

    # Path Initializations
    homeDir = os.path.expanduser('~')
    repoPath = os.path.dirname(os.path.abspath(__file__))
    loc = str(Path(homeDir + '/Downloads/Pageloads.xlsx'))

    # Create empty dictionary for Expected Result
    dict = {}
    
    # Parent Widget
    root = Tk()
    # Read Input File and store expected data in a dictionary as key value pairs
    with open(str(Path(repoPath + '/adobe-analytics-data-raw (16).csv'))) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter='~')
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                print(f'Column names are {", ".join(row)}')
                dict[row[0].strip()]= row[2].strip()
                line_count += 1
            else:
                print(f'\t{row[0]} parameter has {row[2]} value')
                dict[row[0].strip()]= row[2].strip()
                line_count += 1   
        print(f'Processed {line_count} lines.')

    # create workbook and sheet object to refer to Summary sheet
    wb_obj = openpyxl.load_workbook(loc)
    sheet_obj = wb_obj["Summary"]
    
    # Calculate maximum rows and columns
    rows = sheet_obj.max_row
    columns = sheet_obj.max_column
    r = 0
    c = []
    
    for i in range(1,columns + 1):
        cell_obj = sheet_obj.cell(row = 8,column =i)
        if str(dict.get("Current URL")).strip() == str(cell_obj.value).strip() :
            r =  cell_obj.row
            c.append(cell_obj.column)

    rowN = 0
    for keys in dict:
        ttk.Label(root,text=str(keys),width = 20,borderwidth=2, wraplength=100, justify="center", relief="sunken").grid(row=rowN,column=0,columnspan=1,sticky=W)
        ttk.Label(root,text=str(dict.get(keys)).strip(),width =100,borderwidth=2, wraplength=600, justify="center",relief="raised").grid(row=rowN,column=1,columnspan=1,sticky=W)
        flag = False
        for j in range(1,rows + 1):
            for col in c:
               if str(keys).strip() == str(sheet_obj.cell(row = j,column = 1).value).strip():
                    #print(str(keys))
                    #print(j)
                    #print(col)
                     
                    if str(dict.get(keys)).strip() == str(sheet_obj.cell(row = j,column = col).value).strip():
                        #print("Expected value  {} and Actual value  {} matched".format(str(dict.get(keys)),str(sheet_obj.cell(row = j,column = col).value).strip()))
                        #print("Result is pass")
                        ttk.Label(root,text=str(sheet_obj.cell(row = j,column = col).value).strip(),width = 100,borderwidth=2,wraplength=600, justify="center", relief="sunken").grid(row=rowN,column=2,columnspan=1,sticky=W)
                        ttk.Label(root,text="PASS",borderwidth=2, width=20,wraplength=100,relief="raised",justify="center").grid(row=rowN,column=3,columnspan=2,sticky=W)
                        flag = True
                    else:
                        flag = True
                        ttk.Label(root,text=str(sheet_obj.cell(row = j,column = col).value).strip(),width = 100,borderwidth=2,wraplength=600, justify="center", relief="sunken").grid(row=rowN,column=2,columnspan=1,sticky=W)
                        #print("Expected value {} and Actual value {} not matched".format(str(dict.get(keys)),str(sheet_obj.cell(row = j,column = col).value).strip()))
                        #print("Result is fail")
                        ttk.Label(root,text="FAIL",borderwidth=2, width=20,wraplength=100,relief="raised").grid(row=rowN,column=3,columnspan=2,sticky=E)
                break         
        if flag == False:
            #print("key {} not found".format(keys))
            ttk.Label(root,text="None",width = 100,borderwidth=2,wraplength=600, justify="center", relief="sunken").grid(row=rowN,column=2,columnspan=1,sticky=W)
            ttk.Label(root,text="key not found",borderwidth=2, width=20,wraplength=100,relief="raised",justify="center").grid(row=rowN,column=3,columnspan=2,sticky=W)
        rowN = rowN + 1
        root.rowconfigure(rowN, weight=0)



    root.columnconfigure(0, weight=1)
    root.columnconfigure(1, weight=3)
    root.columnconfigure(2, weight=3)
    root.columnconfigure(3, weight=0)
    root.configure(background="white")

    root.mainloop()  


                 


        
        




   





